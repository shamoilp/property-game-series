"""BOE Deal Analyzer - India Edition (NRI Gulf Investor Kit component)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\user\real-estate-game-toolkit\02-nri-gulf-kit\BOE-Deal-Analyzer-India.xlsx"

NAVY = "0E2233"
GOLD = "C9A227"
SAND = "E8DCC0"
PAPER = "FBF8F1"
BLUE = Font(name="Arial", size=10, color="0000FF")          # inputs
BLACK = Font(name="Arial", size=10, color="000000")          # formulas
BOLD = Font(name="Arial", size=10, bold=True)
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SEC = Font(name="Arial", size=10, bold=True, color=NAVY)
TITLE = Font(name="Arial", size=15, bold=True, color=NAVY)
FILL_HDR = PatternFill("solid", start_color=NAVY)
FILL_SEC = PatternFill("solid", start_color=SAND)
FILL_IN = PatternFill("solid", start_color="FFF7DC")
FILL_OUT = PatternFill("solid", start_color=PAPER)
THIN = Border(*[Side(style="thin", color="C9C9C9")] * 4)

wb = Workbook()

# ============================================================ READ ME
rm = wb.active
rm.title = "READ ME"
rm.sheet_view.showGridLines = False
rm.column_dimensions["A"].width = 3
rm.column_dimensions["B"].width = 100
rows = [
    ("T", "BOE DEAL ANALYZER - INDIA EDITION"),
    ("S", "The Property Game Series  |  NRI Gulf Investor Kit"),
    ("", ""),
    ("H", "WHAT THIS DOES"),
    ("B", "A back-of-the-envelope (BOE) check used by professional investors before any commitment: "
          "what the property earns (ROA), what the loan costs (financing constant), and the spread between "
          "them (ROE). If ROA is below your loan cost you have NEGATIVE LEVERAGE - the loan eats your equity "
          "and the deal only works if prices rise. The sheet tells you this in one cell."),
    ("", ""),
    ("H", "HOW TO USE"),
    ("B", "1. Open the 'Deal Analyzer' tab. Enter values ONLY in blue cells (yellow background)."),
    ("B", "2. Compare up to three deals side by side in columns C, D and E."),
    ("B", "3. Read the VERDICT section at the bottom. Green = the rent carries the loan. Red = appreciation bet."),
    ("B", "4. NRI? Fill the NRI ADJUSTMENTS block: TDS on rent and post-tax yield are computed for you."),
    ("", ""),
    ("H", "COLOUR CODE"),
    ("B", "Blue text on light yellow = your inputs.   Black text = calculated, do not type over."),
    ("", ""),
    ("H", "THE FIVE NUMBERS THAT MATTER"),
    ("B", "NOI - net operating income: rent actually kept after vacancy and running costs."),
    ("B", "CFO - cash flow from operations: NOI minus a reserve for repairs and re-letting costs."),
    ("B", "ROA - CFO divided by total purchase cost: the property's own earning power, ignoring the loan."),
    ("B", "Financing constant - annual EMI divided by loan amount: the true annual cost of the debt."),
    ("B", "ROE - cash left after EMI, divided by your own cash in: what YOUR money earns."),
    ("", ""),
    ("H", "RULES OF THUMB (INDIA, 2026)"),
    ("B", "Residential gross yields: 2-3.5% in Tier 1, 3-4.5% in Tier 2/3. Home loans: 8-9.5%."),
    ("B", "That gap means most leveraged residential is negative leverage - know it before you sign."),
    ("B", "Small commercial (shops/offices/clinics): 6-9% yields - can clear loan cost. Run both here."),
    ("", ""),
    ("S", "Educational tool, not investment/tax advice. Verify every number for your own deal."),
]
r = 1
for kind, text in rows:
    c = rm.cell(row=r, column=2, value=text)
    if kind == "T":
        c.font = TITLE
    elif kind == "H":
        c.font = SEC; c.fill = FILL_SEC
    elif kind == "S":
        c.font = Font(name="Arial", size=9, italic=True, color="5A6B7A")
    else:
        c.font = BLACK
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rm.row_dimensions[r].height = 28 if len(text) > 110 else 15
    r += 1

# ============================================================ ANALYZER
ws = wb.create_sheet("Deal Analyzer")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 46
for col in "CDE":
    ws.column_dimensions[col].width = 16
ws.column_dimensions["F"].width = 54

def put(row, label, vals=None, fmla=None, fmt="#,##0", note="", bold=False, inp=False):
    c = ws.cell(row=row, column=2, value=label)
    c.font = BOLD if bold else BLACK
    c.border = THIN
    for i, col in enumerate(("C", "D", "E")):
        cell = ws.cell(row=row, column=3 + i)
        if vals is not None:
            cell.value = vals[i]
        elif fmla is not None:
            cell.value = fmla.format(c=col)
        cell.number_format = fmt
        cell.font = BLUE if inp else BLACK
        cell.fill = FILL_IN if inp else FILL_OUT
        cell.border = THIN
    if note:
        n = ws.cell(row=row, column=6, value=note)
        n.font = Font(name="Arial", size=8.5, color="5A6B7A")
        n.alignment = Alignment(wrap_text=True, vertical="center")

def section(row, text):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value=text)
    c.font = HDR; c.fill = FILL_HDR
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18

ws.cell(row=1, column=2, value="BOE DEAL ANALYZER - INDIA").font = TITLE
ws.cell(row=2, column=2, value="Enter blue cells only. All amounts in INR.").font = Font(
    name="Arial", size=9, italic=True, color="5A6B7A")

hdr = ws.cell(row=4, column=2, value="LINE ITEM"); hdr.font = HDR; hdr.fill = FILL_HDR; hdr.border = THIN
for i, name in enumerate(["DEAL A", "DEAL B", "DEAL C"]):
    c = ws.cell(row=4, column=3 + i, value=name)
    c.font = HDR; c.fill = FILL_HDR; c.border = THIN
    c.alignment = Alignment(horizontal="center")
n = ws.cell(row=4, column=6, value="NOTES"); n.font = HDR; n.fill = FILL_HDR; n.border = THIN

section(5, "STEP 1 - THE PROPERTY")
put(6,  "Purchase price (agreement value)", vals=[9000000, 9000000, 9000000], inp=True)
put(7,  "Stamp duty + registration + GST if any", vals=[630000, 630000, 630000], inp=True,
    note="Typically 6-8% resale; GST 5% applies on under-construction")
put(8,  "Brokerage + initial fit-out", vals=[200000, 200000, 200000], inp=True)
put(9,  "TOTAL PROJECT COST", fmla="=SUM({c}6:{c}8)", bold=True)

section(11, "STEP 2 - WHAT IT EARNS")
put(12, "Expected monthly rent", vals=[22000, 22000, 22000], inp=True,
    note="Use actual quotes from 3 brokers, not the listing price")
put(13, "Vacancy allowance (months/year)", vals=[1, 1, 1], fmt="0.0", inp=True)
put(14, "Society/maintenance charges (annual, owner-paid)", vals=[48000, 48000, 48000], inp=True)
put(15, "Property tax (annual)", vals=[12000, 12000, 12000], inp=True)
put(16, "Insurance + misc (annual)", vals=[6000, 6000, 6000], inp=True)
put(17, "Gross annual rent collected", fmla="={c}12*(12-{c}13)", bold=False)
put(18, "NET OPERATING INCOME (NOI)", fmla="={c}17-{c}14-{c}15-{c}16", bold=True)
put(19, "Repairs + re-letting reserve (% of gross rent)", vals=[0.10, 0.10, 0.10], fmt="0.0%", inp=True,
    note="10% is realistic over a full cycle; 0% is fiction")
put(20, "CASH FLOW FROM OPERATIONS (CFO)", fmla="={c}18-{c}19*{c}17", bold=True)

section(22, "STEP 3 - THE LOAN")
put(23, "Loan amount", vals=[6750000, 6750000, 0], inp=True, note="Banks lend 75-80% LTV to NRIs")
put(24, "Interest rate (annual)", vals=[0.0875, 0.0875, 0.0875], fmt="0.00%", inp=True)
put(25, "Tenure (years)", vals=[20, 20, 20], fmt="0", inp=True)
put(26, "Annual EMI (debt service)",
    fmla="=IF({c}23=0,0,-PMT({c}24/12,{c}25*12,{c}23)*12)", bold=False)
put(27, "Your own cash invested (equity)", fmla="={c}9-{c}23", bold=True)

section(29, "STEP 4 - THE VERDICT")
put(30, "Return on assets (ROA = CFO / total cost)", fmla="={c}20/{c}9", fmt="0.00%", bold=True,
    note="The property's earning power, ignoring the loan")
put(31, "Financing constant (annual EMI / loan)",
    fmla="=IF({c}23=0,0,{c}26/{c}23)", fmt="0.00%", bold=True,
    note="True annual cost of debt incl. principal")
put(32, "Cash flow after financing (CFAF)", fmla="={c}20-{c}26", bold=True)
put(33, "Return on equity (ROE = CFAF / your cash)", fmla="={c}32/{c}27", fmt="0.00%", bold=True)
put(34, "Gross yield (rent / price)", fmla="={c}12*12/{c}6", fmt="0.00%")
put(35, "Net yield (CFO / total cost)", fmla="={c}20/{c}9", fmt="0.00%")
put(36, "LEVERAGE VERDICT",
    fmla='=IF({c}23=0,"UNLEVERAGED",IF({c}30>={c}31,"POSITIVE - rent carries the loan",'
         '"NEGATIVE - appreciation bet"))',
    fmt="General", bold=True,
    note="Negative leverage = every year of flat prices, you lose")
put(37, "Years of EMI your rent cannot cover (cash you feed in/yr)",
    fmla="=MAX(0,-{c}32)", note="If positive, this is your annual top-up from salary")

section(39, "STEP 5 - NRI ADJUSTMENTS (skip if resident)")
put(40, "TDS rate on rent paid to NRI", vals=[0.3120, 0.3120, 0.3120], fmt="0.00%", inp=True,
    note="Sec 195: 30% + cess = 31.2% (surcharge extra at higher slabs); lower via Sec 197 certificate")
put(41, "TDS withheld on rent (annual)", fmla="={c}17*{c}40")
put(42, "Cash reaching you before refund/return filing", fmla="={c}32-{c}41", bold=True,
    note="You may recover excess TDS by filing ITR - but cash timing matters")
put(43, "Post-TDS cash yield on equity", fmla="={c}42/{c}27", fmt="0.00%", bold=True)

# verdict conditional colour
from openpyxl.formatting.rule import FormulaRule
for col in "CDE":
    ws.conditional_formatting.add(
        f"{col}36",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("POSITIVE",{col}36))'],
                    font=Font(color="2E7D5B", bold=True)))
    ws.conditional_formatting.add(
        f"{col}36",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("NEGATIVE",{col}36))'],
                    font=Font(color="B3402E", bold=True)))

ws.freeze_panes = "C5"

# ============================================================ EXAMPLE
ex = wb.create_sheet("Worked Example")
ex.sheet_view.showGridLines = False
ex.column_dimensions["A"].width = 3
ex.column_dimensions["B"].width = 100
example = [
    ("T", "WORKED EXAMPLE - WHY THE VERDICT MATTERS"),
    ("", ""),
    ("B", "Deal A (defaults in the analyzer): a Rs 90 lakh Tier-1 flat renting at Rs 22,000/month, "
          "bought with a 75% loan at 8.75%."),
    ("", ""),
    ("B", "Gross yield looks like 2.9%. After society charges, property tax, one vacant month and a "
          "repairs reserve, CFO is roughly Rs 1.5 lakh - an ROA of about 1.5%."),
    ("B", "The loan's financing constant is about 10.6% (EMI of ~Rs 7.16 lakh on a Rs 67.5 lakh loan)."),
    ("B", "1.5% earned vs 10.6% paid = deeply NEGATIVE leverage. The sheet shows you feed the flat "
          "roughly Rs 5.6 lakh per year from salary. Over 5 flat-price years that is Rs 28 lakh of top-ups."),
    ("", ""),
    ("B", "Now set Deal C's loan to zero (already done): same flat, all cash. ROE = ROA = ~1.5%. "
          "Still weak - but at least nothing is eating your equity. The flat must appreciate ~7%+ a year "
          "just to beat an FD after costs."),
    ("", ""),
    ("B", "Change Deal B to a small commercial unit: Rs 90 lakh shop renting at Rs 55,000/month. "
          "ROA jumps to ~6.3%; against a 9.5% LAP-style loan it is still negative - but with 40% loan "
          "it turns positive. This is the game: the spread, not the asset class, decides."),
    ("", ""),
    ("S", "Replace every default with your actual deal. The verdict cell does the rest."),
]
r = 1
for kind, text in example:
    c = ex.cell(row=r, column=2, value=text)
    if kind == "T":
        c.font = TITLE
    elif kind == "S":
        c.font = Font(name="Arial", size=9, italic=True, color="5A6B7A")
    else:
        c.font = BLACK
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ex.row_dimensions[r].height = 30 if len(text) > 100 else 15
    r += 1

wb.save(OUT)
print("Saved:", OUT)
