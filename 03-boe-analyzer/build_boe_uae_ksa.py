"""BOE Deal Analyzer - UAE and KSA editions. Same engine as the India
edition, market-specific cost stacks and financing blocks."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

DIR = r"C:\Users\user\real-estate-game-toolkit\03-boe-analyzer"
NAVY, SAND, PAPER = "0E2233", "E8DCC0", "FBF8F1"
BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
BOLD = Font(name="Arial", size=10, bold=True)
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SEC = Font(name="Arial", size=10, bold=True, color=NAVY)
TITLE = Font(name="Arial", size=15, bold=True, color=NAVY)
NOTE = Font(name="Arial", size=8.5, color="5A6B7A")
ITAL = Font(name="Arial", size=9, italic=True, color="5A6B7A")
FILL_HDR = PatternFill("solid", start_color=NAVY)
FILL_SEC = PatternFill("solid", start_color=SAND)
FILL_IN = PatternFill("solid", start_color="FFF7DC")
FILL_OUT = PatternFill("solid", start_color=PAPER)
THIN = Border(*[Side(style="thin", color="C9C9C9")] * 4)


def readme(wb, title, lines):
    rm = wb.active
    rm.title = "READ ME"
    rm.sheet_view.showGridLines = False
    rm.column_dimensions["A"].width = 3
    rm.column_dimensions["B"].width = 100
    r = 1
    for kind, text in [("T", title)] + lines:
        c = rm.cell(row=r, column=2, value=text)
        if kind == "T":
            c.font = TITLE
        elif kind == "H":
            c.font = SEC; c.fill = FILL_SEC
        elif kind == "S":
            c.font = ITAL
        else:
            c.font = BLACK
            c.alignment = Alignment(wrap_text=True, vertical="top")
            rm.row_dimensions[r].height = 28 if len(text) > 110 else 15
        r += 1


def analyzer_sheet(wb, sheet_title, currency, rows_spec, verdict_row, deal_names):
    ws = wb.create_sheet("Deal Analyzer")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 47
    for col in "CDE":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].width = 54

    ws.cell(row=1, column=2, value=sheet_title).font = TITLE
    ws.cell(row=2, column=2,
            value=f"Enter blue cells only. All amounts in {currency}.").font = ITAL
    h = ws.cell(row=4, column=2, value="LINE ITEM")
    h.font = HDR; h.fill = FILL_HDR; h.border = THIN
    for i, name in enumerate(deal_names):
        c = ws.cell(row=4, column=3 + i, value=name)
        c.font = HDR; c.fill = FILL_HDR; c.border = THIN
        c.alignment = Alignment(horizontal="center")
    n = ws.cell(row=4, column=6, value="NOTES")
    n.font = HDR; n.fill = FILL_HDR; n.border = THIN

    for spec in rows_spec:
        kind = spec[0]
        if kind == "sec":
            _, row, text = spec
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            c = ws.cell(row=row, column=2, value=text)
            c.font = HDR; c.fill = FILL_HDR
            c.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 18
        else:
            _, row, label, vals, fmla, fmt, note, bold, inp = spec
            c = ws.cell(row=row, column=2, value=label)
            c.font = BOLD if bold else BLACK
            c.border = THIN
            for i, col in enumerate(("C", "D", "E")):
                cell = ws.cell(row=row, column=3 + i)
                if vals is not None:
                    cell.value = vals[i]
                elif fmla is not None:
                    cell.value = fmla.format(c=col)
                cell.number_format = fmt
                cell.font = BLUE if inp else BLACK
                cell.fill = FILL_IN if inp else FILL_OUT
                cell.border = THIN
            if note:
                nn = ws.cell(row=row, column=6, value=note)
                nn.font = NOTE
                nn.alignment = Alignment(wrap_text=True, vertical="center")

    for col in "CDE":
        ws.conditional_formatting.add(
            f"{col}{verdict_row}",
            FormulaRule(formula=[f'ISNUMBER(SEARCH("POSITIVE",{col}{verdict_row}))'],
                        font=Font(color="2E7D5B", bold=True)))
        ws.conditional_formatting.add(
            f"{col}{verdict_row}",
            FormulaRule(formula=[f'ISNUMBER(SEARCH("NEGATIVE",{col}{verdict_row}))'],
                        font=Font(color="B3402E", bold=True)))
    ws.freeze_panes = "C5"


def row(r, label, vals=None, fmla=None, fmt="#,##0", note="", bold=False, inp=False):
    return ("row", r, label, vals, fmla, fmt, note, bold, inp)


def sec(r, text):
    return ("sec", r, text)


# =====================================================================
# UAE EDITION
# =====================================================================
wb = Workbook()
readme(wb, "BOE DEAL ANALYZER - UAE EDITION", [
    ("S", "The Property Game Series"),
    ("", ""),
    ("H", "WHAT THIS DOES"),
    ("B", "The professional pre-commitment check for a UAE buy-to-let: what the unit earns after "
          "the REAL cost stack (service charges, chiller, management), what the mortgage costs, and "
          "whether the rent carries the loan. Headline yields in the UAE die by service charge - "
          "this sheet makes that visible before you sign the MOU."),
    ("", ""),
    ("H", "HOW TO USE"),
    ("B", "1. Blue cells (yellow background) are inputs. Compare three deals in columns C-E."),
    ("B", "2. Rents are annual (UAE convention). Service charges are AED per sq ft per year - "
          "pull the actual figure from the Dubai REST app / Mollak record, never the broker's estimate."),
    ("B", "3. Read the VERDICT row. Green = rent carries the mortgage. Red = appreciation bet."),
    ("", ""),
    ("H", "UAE-SPECIFIC TRAPS THIS SHEET FORCES YOU TO PRICE"),
    ("B", "DLD transfer fee 4% + trustee/admin fees: your cost basis is not the listing price."),
    ("B", "Service charges of AED 12-30+/sqft/yr can consume 25-45% of gross rent in some towers."),
    ("B", "Chiller (district cooling) capacity charges sometimes sit with the owner even when vacant."),
    ("B", "Off-plan: a 'post-handover payment plan' is a loan from the developer - enter its instalments "
          "as financing to see the true leverage."),
    ("B", "5% VAT applies to COMMERCIAL rent (you must register above threshold); residential rent is exempt."),
    ("", ""),
    ("S", "Educational tool, not investment advice. Rates and fees as of June 2026 - verify current."),
])
analyzer_sheet(
    wb, "BOE DEAL ANALYZER - UAE", "AED",
    [
        sec(5, "STEP 1 - THE PROPERTY"),
        row(6, "Purchase price", vals=[1500000, 1500000, 1500000], inp=True),
        row(7, "DLD transfer fee + trustee + admin", fmla="={c}6*0.04+5000",
            note="4% DLD + ~AED 4-5k trustee/admin (Dubai); other emirates differ"),
        row(8, "Agency fee + initial fit-out", vals=[40000, 40000, 40000], inp=True,
            note="Agency typically 2% + VAT"),
        row(9, "TOTAL PROJECT COST", fmla="=SUM({c}6:{c}8)", bold=True),
        sec(11, "STEP 2 - WHAT IT EARNS"),
        row(12, "Expected ANNUAL rent", vals=[95000, 95000, 95000], inp=True,
            note="Check the DLD rental index / actual Ejari registrations for the tower"),
        row(13, "Vacancy allowance (weeks/year)", vals=[4, 4, 4], fmt="0", inp=True),
        row(14, "Unit area (sq ft)", vals=[850, 850, 850], fmt="#,##0", inp=True),
        row(15, "Service charge rate (AED/sqft/year)", vals=[18, 18, 18], fmt="0.00", inp=True,
            note="Verify on Mollak/REST app - the single most lied-about number in UAE listings"),
        row(16, "Service charges (annual)", fmla="={c}14*{c}15"),
        row(17, "Chiller/utilities owner-borne (annual)", vals=[4000, 4000, 4000], inp=True),
        row(18, "Property management fee (% of rent)", vals=[0.05, 0.05, 0.05], fmt="0.0%", inp=True),
        row(19, "Gross annual rent collected", fmla="={c}12*(52-{c}13)/52"),
        row(20, "NET OPERATING INCOME (NOI)",
            fmla="={c}19-{c}16-{c}17-{c}18*{c}19", bold=True),
        row(21, "Repairs + re-letting reserve (% of gross)", vals=[0.07, 0.07, 0.07], fmt="0.0%", inp=True),
        row(22, "CASH FLOW FROM OPERATIONS (CFO)", fmla="={c}20-{c}21*{c}19", bold=True),
        sec(24, "STEP 3 - THE MORTGAGE (or developer payment plan)"),
        row(25, "Loan amount", vals=[1125000, 1125000, 0], inp=True,
            note="Expat LTV cap commonly 80% (first home, ready). Off-plan: enter plan instalments' financed portion"),
        row(26, "Interest/profit rate (annual)", vals=[0.045, 0.045, 0.045], fmt="0.00%", inp=True),
        row(27, "Tenure (years)", vals=[25, 25, 25], fmt="0", inp=True),
        row(28, "Annual debt service",
            fmla="=IF({c}25=0,0,-PMT({c}26/12,{c}27*12,{c}25)*12)"),
        row(29, "Your own cash invested (equity)", fmla="={c}9-{c}25", bold=True),
        sec(31, "STEP 4 - THE VERDICT"),
        row(32, "Return on assets (ROA = CFO / total cost)", fmla="={c}22/{c}9", fmt="0.00%", bold=True),
        row(33, "Financing constant (debt service / loan)",
            fmla="=IF({c}25=0,0,{c}28/{c}25)", fmt="0.00%", bold=True),
        row(34, "Cash flow after financing (CFAF)", fmla="={c}22-{c}28", bold=True),
        row(35, "Return on equity (ROE)", fmla="={c}34/{c}29", fmt="0.00%", bold=True),
        row(36, "Gross yield (advertised look)", fmla="={c}12/{c}6", fmt="0.00%",
            note="The number the broker quotes"),
        row(37, "TRUE net yield (CFO / total cost)", fmla="={c}22/{c}9", fmt="0.00%",
            note="The number your bank account experiences"),
        row(38, "LEVERAGE VERDICT",
            fmla='=IF({c}25=0,"UNLEVERAGED",IF({c}32>={c}33,'
                 '"POSITIVE - rent carries the loan","NEGATIVE - appreciation bet"))',
            fmt="General", bold=True),
        row(39, "Annual top-up if rent falls short", fmla="=MAX(0,-{c}34)"),
        sec(41, "STEP 5 - SERVICE CHARGE STRESS TEST"),
        row(42, "Service charge share of gross rent", fmla="={c}16/{c}19", fmt="0.0%", bold=True,
            note="Above 30% = the tower works for the OA, not for you"),
        row(43, "Net yield if service charge rises 25%",
            fmla="=({c}22-0.25*{c}16)/{c}9", fmt="0.00%",
            note="Service charges only ever move one way"),
    ],
    verdict_row=38,
    deal_names=["DEAL A", "DEAL B", "DEAL C"],
)
wb.save(rf"{DIR}\BOE-Deal-Analyzer-UAE.xlsx")
print("Saved UAE edition")

# =====================================================================
# KSA EDITION
# =====================================================================
wb = Workbook()
readme(wb, "BOE DEAL ANALYZER - KSA EDITION", [
    ("S", "The Property Game Series  |  Shariah-aware financing block"),
    ("", ""),
    ("H", "WHAT THIS DOES"),
    ("B", "The pre-commitment check for a Saudi buy-to-let, built for how KSA actually works: "
          "Murabaha/Ijara financing instead of a conventional mortgage, Ejar-registered leases, RETT "
          "on transfer, and a Zakat awareness block. It answers one question: does the rent carry the "
          "financing, or are you betting on Vision-2030 appreciation?"),
    ("", ""),
    ("H", "HOW TO USE"),
    ("B", "1. Blue cells (yellow background) are inputs. Compare three deals in columns C-E."),
    ("B", "2. Rents are annual (KSA convention). Pull comparable rents from the Ejar index for the district."),
    ("B", "3. The financing block uses the bank's quoted ANNUAL PROFIT RATE (APR) on Murabaha/Ijara - "
          "ask for the APR, not the flat rate. A '3.5% flat' is roughly 6.5%+ APR."),
    ("", ""),
    ("H", "KSA-SPECIFIC ITEMS THIS SHEET PRICES"),
    ("B", "RETT 5% on transfer value - confirm who bears it in your deal; exemptions exist (e.g., first-home "
          "support up to a cap)."),
    ("B", "Ejar registration makes your lease enforceable and eviction process defined - never rent outside it."),
    ("B", "White Land Levy: vacant urban land can attract an annual levy (regime expanded in 2025 - verify the "
          "current rate and whether your plot is in scope). Enter it if you are holding land."),
    ("B", "Zakat block: property bought WITH INTENT TO RESELL is Zakatable at ~2.577% of market value per "
          "Hijri year (2.5% adjusted to the solar year). Rental property: the accumulated net rent is "
          "Zakatable with your cash. Record your intention at purchase."),
    ("", ""),
    ("S", "Educational tool, not investment, tax or Shariah advice. Verify rates - June 2026."),
])
analyzer_sheet(
    wb, "BOE DEAL ANALYZER - KSA", "SAR",
    [
        sec(5, "STEP 1 - THE PROPERTY"),
        row(6, "Purchase price", vals=[1200000, 1200000, 1200000], inp=True),
        row(7, "RETT borne by you (5% unless exempt/shifted)", fmla="={c}6*0.05",
            note="Legally on the seller-disponor side but commercially negotiated; first-home exemption may apply"),
        row(8, "Broker fee (max 2.5%) + initial works", vals=[40000, 40000, 40000], inp=True),
        row(9, "TOTAL PROJECT COST", fmla="=SUM({c}6:{c}8)", bold=True),
        sec(11, "STEP 2 - WHAT IT EARNS"),
        row(12, "Expected ANNUAL rent (Ejar comparable)", vals=[65000, 65000, 65000], inp=True,
            note="Use Ejar index for the district, not the asking rents on listing apps"),
        row(13, "Vacancy allowance (weeks/year)", vals=[4, 4, 4], fmt="0", inp=True),
        row(14, "Maintenance + repairs (annual)", vals=[8000, 8000, 8000], inp=True),
        row(15, "Management/wakeel fee (% of rent)", vals=[0.05, 0.05, 0.05], fmt="0.0%", inp=True),
        row(16, "Other owner costs (insurance, municipality)", vals=[3000, 3000, 3000], inp=True),
        row(17, "Gross annual rent collected", fmla="={c}12*(52-{c}13)/52"),
        row(18, "NET OPERATING INCOME (NOI)",
            fmla="={c}17-{c}14-{c}15*{c}17-{c}16", bold=True),
        row(19, "Re-letting reserve (% of gross)", vals=[0.05, 0.05, 0.05], fmt="0.0%", inp=True),
        row(20, "CASH FLOW FROM OPERATIONS (CFO)", fmla="={c}18-{c}19*{c}17", bold=True),
        sec(22, "STEP 3 - ISLAMIC FINANCING (Murabaha / Ijara)"),
        row(23, "Financed amount", vals=[840000, 840000, 0], inp=True,
            note="REDF/Sakani support can change this block entirely - model with and without"),
        row(24, "Annual profit rate - APR equivalent", vals=[0.065, 0.065, 0.065], fmt="0.00%", inp=True,
            note="Insist on APR. Flat rates roughly double when converted"),
        row(25, "Tenure (years)", vals=[20, 20, 20], fmt="0", inp=True),
        row(26, "Annual instalments",
            fmla="=IF({c}23=0,0,-PMT({c}24/12,{c}25*12,{c}23)*12)"),
        row(27, "Your own cash invested (equity)", fmla="={c}9-{c}23", bold=True),
        sec(29, "STEP 4 - THE VERDICT"),
        row(30, "Return on assets (ROA = CFO / total cost)", fmla="={c}20/{c}9", fmt="0.00%", bold=True),
        row(31, "Financing constant (instalments / financed amt)",
            fmla="=IF({c}23=0,0,{c}26/{c}23)", fmt="0.00%", bold=True),
        row(32, "Cash flow after financing (CFAF)", fmla="={c}20-{c}26", bold=True),
        row(33, "Return on equity (ROE)", fmla="={c}32/{c}27", fmt="0.00%", bold=True),
        row(34, "Gross yield", fmla="={c}12/{c}6", fmt="0.00%"),
        row(35, "TRUE net yield (CFO / total cost)", fmla="={c}20/{c}9", fmt="0.00%"),
        row(36, "LEVERAGE VERDICT",
            fmla='=IF({c}23=0,"UNLEVERAGED",IF({c}30>={c}31,'
                 '"POSITIVE - rent carries the financing","NEGATIVE - appreciation bet"))',
            fmt="General", bold=True),
        row(37, "Annual top-up if rent falls short", fmla="=MAX(0,-{c}32)"),
        sec(39, "STEP 5 - ZAKAT AWARENESS (record your intention)"),
        row(40, "Intention at purchase: 1 = resale/trading, 2 = hold for rent",
            vals=[2, 2, 1], fmt="0", inp=True,
            note="Niyyah drives Zakat treatment - decide and record it on day one"),
        row(41, "Estimated current market value", vals=[1200000, 1200000, 1200000], inp=True),
        row(42, "Indicative annual Zakat if TRADING intent (2.577% of value)",
            fmla='=IF({c}40=1,{c}41*0.02577,0)', bold=True,
            note="Hold-for-rent: property itself not Zakatable; your saved net rent is, with your cash"),
        row(43, "CFO after indicative Zakat", fmla="={c}20-{c}42"),
        sec(45, "STEP 6 - LAND HOLDING ONLY"),
        row(46, "Vacant urban land? Annual White Land Levy estimate", vals=[0, 0, 0], inp=True,
            note="Expanded regime since 2025 - rates and scope vary; check your parcel's status before buying land to bank"),
        row(47, "CFO after land levy", fmla="={c}20-{c}46"),
    ],
    verdict_row=36,
    deal_names=["DEAL A", "DEAL B", "DEAL C"],
)
wb.save(rf"{DIR}\BOE-Deal-Analyzer-KSA.xlsx")
print("Saved KSA edition")
